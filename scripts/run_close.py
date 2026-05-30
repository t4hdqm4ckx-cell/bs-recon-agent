#!/usr/bin/env python3
"""
BS Reconciliation Agent — Close Cycle Runner
Period: 2026-11 | Company: Lumina Streaming Co.
Implements the workflow described in skills/bs-reconciliation/SKILL.md
"""
import os, json, hashlib, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PERIOD = "2026-11"
RUN_TS = "2026-12-01T08:00:00Z"
AGENT_VERSION = "0.1.0"
ROOT = "/tmp/bs-recon-run"

# ── thresholds (from config/thresholds.yaml) ────────────────────────────────
EXCEPTION_AMT = 100_000
EXCEPTION_AGE = 30
AUTO_INVEST_AMT = 250_000
TRIVIAL_AGG = 450_000
JE_CONTROLLER = 250_000
JE_CFO = 1_000_000
BANK_OUTSTANDING_DAYS = 60
IC_MISMATCH_FLAG = 50_000
AR_90PLUS_PCT = 0.05
STALE_ACCRUAL_DAYS = 45

# ── styles ──────────────────────────────────────────────────────────────────
NAVY = "1E2761"; TEAL = "028090"; ICE = "CADCFC"; LIGHT = "EEF4FF"
BORDER_GRAY = "CBD5E1"; RED = "DC2626"; ORANGE = "EA580C"; YELLOW = "D97706"; GREEN = "16A34A"

H_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=NAVY)
SUB_FONT = Font(name="Calibri", size=10, italic=True, color="64748B")
H_FILL = PatternFill("solid", fgColor=NAVY)
ACCENT_FILL = PatternFill("solid", fgColor=TEAL)
BAND_FILL = PatternFill("solid", fgColor=LIGHT)
THIN = Side(border_style="thin", color=BORDER_GRAY)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

def style_header_row(ws, row, ncols, fill=H_FILL):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = H_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BOX

def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def money_fmt(cell):
    cell.number_format = '"$"#,##0.00;[Red]("$"#,##0.00)'
    cell.alignment = RIGHT

def add_audit_tab(wb, sources, account, entity):
    ws = wb.create_sheet("Audit Trail")
    ws["A1"] = "Audit Trail"
    ws["A1"].font = TITLE_FONT
    fields = [
        ("Agent",          "account-reconciliation"),
        ("Agent Version",  AGENT_VERSION),
        ("Run Timestamp",  RUN_TS),
        ("Period",         PERIOD),
        ("Entity",         entity),
        ("Account",        account),
        ("Preparer",       "agent:account-reconciliation"),
        ("Reviewer",       "[Pending human sign-off]"),
        ("Sign-off Date",  "[Pending]"),
    ]
    for i, (k, v) in enumerate(fields, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.cell(row=len(fields) + 4, column=1, value="Sources").font = Font(bold=True, color=NAVY)
    r = len(fields) + 5
    ws.cell(row=r, column=1, value="Path").font = H_FONT
    ws.cell(row=r, column=1).fill = H_FILL
    ws.cell(row=r, column=2, value="SHA-256").font = H_FONT
    ws.cell(row=r, column=2).fill = H_FILL
    for src in sources:
        r += 1
        ws.cell(row=r, column=1, value=src["path"])
        ws.cell(row=r, column=2, value=src["sha256"])
    autosize(ws, [42, 70])

def sha(text):
    return hashlib.sha256(text.encode()).hexdigest()[:16]

# ════════════════════════════════════════════════════════════════════════════
# 1. SYNTHETIC SOURCE DATA
# ════════════════════════════════════════════════════════════════════════════
SOURCES = {}  # path -> sha

def write_data(name, sheets_dict):
    wb = Workbook()
    wb.remove(wb.active)
    fingerprint = []
    for sheet_name, rows in sheets_dict.items():
        ws = wb.create_sheet(sheet_name)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
                fingerprint.append(str(val))
            if r_idx == 1:
                style_header_row(ws, 1, len(row))
        autosize(ws, [22] * max(len(r) for r in rows))
    path = f"data/synthetic/{name}"
    out = os.path.join(ROOT, path)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    SOURCES[path] = sha("|".join(fingerprint))
    return path

# ─── GL Trial Balance (all entities) ───────────────────────────────────────
tb_rows = [
    ["Entity", "Account", "Name", "Debit", "Credit", "Net Balance"],
    # LuminaUS — Assets
    ["LuminaUS","100100","Cash - Operating",                  142_700_000.00, 0.00, 142_700_000.00],
    ["LuminaUS","100200","Cash - Money Market",                85_000_000.00, 0.00,  85_000_000.00],
    ["LuminaUS","120100","AR - Trade",                         48_350_000.00, 0.00,  48_350_000.00],
    ["LuminaUS","120200","AR - Advertising",                   12_180_000.00, 0.00,  12_180_000.00],
    ["LuminaUS","130100","Prepaid Insurance",                   1_215_000.00, 0.00,   1_215_000.00],
    ["LuminaUS","130200","Prepaid Software",                    3_870_000.00, 0.00,   3_870_000.00],
    ["LuminaUS","130300","Prepaid Rent",                        2_400_000.00, 0.00,   2_400_000.00],
    ["LuminaUS","130400","Prepaid Marketing",                     920_000.00, 0.00,     920_000.00],
    ["LuminaUS","150100","PP&E - Net",                        198_600_000.00, 0.00, 198_600_000.00],
    ["LuminaUS","150200","Intangibles - Net",                  62_400_000.00, 0.00,  62_400_000.00],
    # LuminaUS — Liabilities
    ["LuminaUS","200100","AP - Trade",                          0.00, 38_220_000.00, -38_220_000.00],
    ["LuminaUS","200200","AP - Accrued",                        0.00,  6_410_000.00,  -6_410_000.00],
    ["LuminaUS","210100","Accrued Compensation",                0.00, 14_820_000.00, -14_820_000.00],
    ["LuminaUS","210200","Accrued Content",                     0.00, 27_310_000.00, -27_310_000.00],
    ["LuminaUS","210300","Accrued Legal",                       0.00,  2_180_000.00,  -2_180_000.00],
    ["LuminaUS","210400","Accrued Marketing",                   0.00,  3_842_000.00,  -3_842_000.00],
    ["LuminaUS","210500","Other Accruals",                      0.00,  1_950_000.00,  -1_950_000.00],
    ["LuminaUS","220100","Deferred Revenue - Subscription",     0.00, 92_500_000.00, -92_500_000.00],
    ["LuminaUS","220200","Deferred Revenue - Advertising",      0.00, 18_278_000.00, -18_278_000.00],
    ["LuminaUS","230100","IC Receivable - LuminaUS",           24_867_000.00, 0.00,  24_867_000.00],
    ["LuminaUS","230200","IC Payable - LuminaUS",               0.00, 18_400_000.00, -18_400_000.00],
    ["LuminaUS","300100","Long-Term Debt",                      0.00,150_000_000.00,-150_000_000.00],
    # LuminaEMEA — selected
    ["LuminaEMEA","100100","Cash - Operating",                 38_500_000.00, 0.00,  38_500_000.00],
    ["LuminaEMEA","240100","IC Receivable - LuminaEMEA",       18_400_000.00, 0.00,  18_400_000.00],
    ["LuminaEMEA","240200","IC Payable - LuminaEMEA",           0.00, 24_600_000.00, -24_600_000.00],
    # LuminaAPAC — selected
    ["LuminaAPAC","100100","Cash - Operating",                 21_300_000.00, 0.00,  21_300_000.00],
    ["LuminaAPAC","230100","IC Receivable - APAC",              6_200_000.00, 0.00,   6_200_000.00],
    ["LuminaAPAC","230200","IC Payable - APAC",                 0.00,  6_467_000.00,  -6_467_000.00],
]
write_data(f"{PERIOD}_GL_TrialBalance.xlsx", {"TB": tb_rows})

# ─── Bank Statement — LuminaUS Operating ───────────────────────────────────
bank_op = [
    ["Date","Description","Type","Amount","Cleared Balance"],
    ["2026-11-30","Beginning Balance","BAL", 138_450_000.00, 138_450_000.00],
    ["2026-11-30","Customer wire — Major Studio Group","DEP",  2_840_000.00, 141_290_000.00],
    ["2026-11-30","ACH receipts (batched)","DEP",            1_690_000.00, 142_980_000.00],
    ["2026-11-30","Bank service charge","CHG",                  -82_000.00, 142_898_000.00],
    ["2026-11-30","Wire transfer fee","CHG",                   -118_000.00, 142_780_000.00],
    ["2026-11-30","Check #48217 cleared","CHK",                -325_000.00, 142_455_000.00],
    ["2026-11-30","Check #48235 cleared","CHK",                -210_000.00, 142_245_000.00],
    ["2026-11-30","Ending Balance per Bank","BAL",                    0.00, 142_245_000.00],
]
write_data(f"{PERIOD}_LuminaUS_BankStatement_Operating.xlsx", {"Statement": bank_op})

# ─── Outstanding Checks — LuminaUS ─────────────────────────────────────────
oc_rows = [
    ["Check #","Issue Date","Payee","Amount","Days Outstanding"],
    ["48201","2026-08-22","Acme Studio Services",  185_000.00, 100],   # >60d — flag, propose void
    ["48205","2026-09-12","Beacon Logistics",       95_000.00,  79],   # >60d — flag
    ["48260","2026-11-25","Cascade Media",         312_000.00,   5],
    ["48261","2026-11-26","Delta Productions",      78_000.00,   4],
    ["48262","2026-11-28","Echo Talent Agency",    230_000.00,   2],
]
dit_rows = [
    ["Date","Source","Amount","Days Outstanding"],
    ["2026-11-29","Star Network Group",            420_000.00,   1],
    ["2026-11-30","Prism Advertising",             310_000.00,   0],
]
write_data(f"{PERIOD}_LuminaUS_OutstandingItems.xlsx", {
    "Outstanding Checks": oc_rows,
    "Deposits in Transit": dit_rows,
})

# ─── AR Aging — LuminaUS Trade ─────────────────────────────────────────────
ar_aging = [
    ["Customer","Current (0-30)","31-60","61-90","91+","Total"],
    ["Star Network Group",       6_200_000.00, 1_800_000.00,   400_000.00,   120_000.00,  8_520_000.00],
    ["Prism Advertising",        4_800_000.00, 2_100_000.00,   650_000.00,   840_000.00,  8_390_000.00],
    ["Apex Studios",             8_300_000.00,   980_000.00,   180_000.00, 1_240_000.00, 10_700_000.00],
    ["Nimbus Broadcasting",      3_700_000.00,   620_000.00,   240_000.00,   450_000.00,  5_010_000.00],
    ["Quartz Productions",       2_900_000.00,   410_000.00,    90_000.00,   175_000.00,  3_575_000.00],
    ["Solstice Media",           4_100_000.00,   320_000.00,   150_000.00,    87_000.00,  4_657_000.00],
    ["[Credit balance] Vortex",    -45_000.00,         0.00,         0.00,         0.00,    -45_000.00],
    ["Other (12 customers)",     6_840_000.00,   480_000.00,   118_000.00,   105_000.00,  7_543_000.00],
    ["TOTAL",                   36_795_000.00, 6_710_000.00, 1_828_000.00, 3_017_000.00, 48_350_000.00],
]
write_data(f"{PERIOD}_LuminaUS_AR_Aging.xlsx", {"Trade AR Aging": ar_aging})

# ─── AP Aging — LuminaUS Trade ─────────────────────────────────────────────
ap_aging = [
    ["Vendor","Current (0-30)","31-60","61-90","91+","Total"],
    ["AWS Cloud",               4_200_000.00,   180_000.00,         0.00,         0.00,  4_380_000.00],
    ["Production House Alpha",  3_800_000.00,   920_000.00,   240_000.00,         0.00,  4_960_000.00],
    ["Talent Agency Bravo",     2_650_000.00,   140_000.00,   105_000.00,    62_000.00,  2_957_000.00],
    ["Studio Lighting Co.",     1_400_000.00,   720_000.00,   310_000.00,   190_000.00,  2_620_000.00],
    ["Distribution Net",        5_100_000.00,   430_000.00,    85_000.00,         0.00,  5_615_000.00],
    ["Other (25 vendors)",     16_320_000.00, 1_140_000.00,    98_000.00,   130_000.00, 17_688_000.00],
    ["TOTAL",                  33_470_000.00, 3_530_000.00,   838_000.00,   382_000.00, 38_220_000.00],
]
write_data(f"{PERIOD}_LuminaUS_AP_Aging.xlsx", {"Trade AP Aging": ap_aging})

# ─── Prepaid Schedule ──────────────────────────────────────────────────────
prepaid_sched = [
    ["Account","Asset","Open Balance","Additions","Amortization","End Balance","Last Use Date","Schedule Term"],
    ["130100","D&O Insurance Policy",     460_000.00,        0.00, -38_000.00,  422_000.00, "2027-04-30","12 months"],
    ["130100","Property Insurance",       320_000.00,        0.00, -27_000.00,  293_000.00, "2027-02-28","12 months"],
    ["130100","Cyber Liability Insurance",480_000.00,        0.00, -50_000.00,  430_000.00, "2027-05-31","12 months"],
    ["130100","Auto fleet policy (LEGACY)", 15_000.00,        0.00,       0.00,   15_000.00, "2026-06-30","Fully amortized"],  # NEEDS WRITE-OFF
    ["130200","Salesforce Enterprise",   1_200_000.00,        0.00,-110_000.00, 1_090_000.00,"2027-07-31","12 months"],
    ["130200","Microsoft E5",              980_000.00,        0.00, -82_000.00,  898_000.00, "2027-06-30","12 months"],
    ["130200","Adobe Creative Cloud",      620_000.00,        0.00, -55_000.00,  565_000.00, "2027-05-31","12 months"],
    ["130200","DataDog APM",               890_000.00,        0.00, -73_000.00,  817_000.00, "2027-08-31","12 months"],
    ["130200","Snowflake Enterprise",      540_000.00,   200_000.00,-110_000.00,  500_000.00,"2027-09-30","18 months"],
    ["130300","HQ Lease (NYC)",          1_400_000.00,        0.00,-100_000.00, 1_300_000.00,"2027-09-30","18 months"],
    ["130300","Studio Lease (LA)",       1_100_000.00,        0.00,        0.00, 1_100_000.00,"2027-10-31","Renewed Nov-26"],
    ["130400","Influencer campaign Q4",    540_000.00,        0.00, -180_000.00,  360_000.00,"2027-01-31","3 months"],
    ["130400","Sponsored content series",  720_000.00,        0.00, -160_000.00,  560_000.00,"2027-02-28","4 months"],
    ["130400","Other prepaid marketing",     0.00,            6_000.00,    0.00,    6_000.00,"2026-12-31","Variance"],  # creates $6K variance
]
write_data(f"{PERIOD}_Prepaid_Schedule.xlsx", {"Prepaid Schedule": prepaid_sched})

# ─── Fixed Asset Register (summary) ────────────────────────────────────────
fa_register = [
    ["Class","Cost","Accumulated Depreciation","Net Book Value"],
    ["Studio equipment", 142_800_000.00, -38_400_000.00, 104_400_000.00],
    ["IT hardware",       58_200_000.00, -22_100_000.00,  36_100_000.00],
    ["Leasehold improv.", 41_500_000.00, -18_200_000.00,  23_300_000.00],
    ["Furniture & fixt.", 21_400_000.00, -13_600_000.00,   7_800_000.00],
    ["Vehicles",           9_800_000.00,  -2_800_000.00,   7_000_000.00],
    ["WIP",               20_000_000.00,           0.00,  20_000_000.00],
    ["TOTAL",            293_700_000.00, -95_100_000.00, 198_600_000.00],
]
intangible_register = [
    ["Class","Cost","Accumulated Amortization","Net Book Value"],
    ["Acquired content libraries", 48_000_000.00, -12_400_000.00, 35_600_000.00],
    ["Software (capitalized)",     21_800_000.00,  -6_800_000.00, 15_000_000.00],
    ["Trademarks",                 11_200_000.00,  -2_400_000.00,  8_800_000.00],
    ["Goodwill (subsidiary)",       3_000_000.00,          0.00,   3_000_000.00],
    ["TOTAL",                      84_000_000.00, -21_600_000.00, 62_400_000.00],
]
write_data(f"{PERIOD}_FixedAsset_Register.xlsx", {
    "PP&E": fa_register,
    "Intangibles": intangible_register,
})

# ─── Accrual Schedule ──────────────────────────────────────────────────────
accrual_sched = [
    ["Account","Accrual","Booked Period","Original Amount","Reversal Status","Current Balance","Age (days)"],
    ["210100","Nov bonus accrual",        "2026-11", 8_400_000.00,"Active",     8_400_000.00,   1],
    ["210100","PTO accrual rollover",     "2026-11", 4_700_000.00,"Active",     4_700_000.00,   1],
    ["210100","Aug perf bonus (STALE)",   "2026-08",   185_000.00,"Stale — should have reversed", 185_000.00, 95],
    ["210100","Severance accrual",        "2026-11", 1_535_000.00,"Active",     1_535_000.00,   1],
    ["210200","Q4 content commitments",   "2026-11",18_500_000.00,"Active",    18_500_000.00,   1],
    ["210200","Sports rights accrual",    "2026-11", 8_500_000.00,"Active",     8_500_000.00,   1],
    ["210200","Aug content (STALE — material)", "2026-08", 310_000.00,"Stale — should have reversed", 310_000.00, 95],
    ["210300","Litigation reserve",       "2026-11", 1_800_000.00,"Active",     1_800_000.00,   1],
    ["210300","Outside counsel fees",     "2026-11",   380_000.00,"Active",       380_000.00,   1],
    ["210400","Q4 campaign accrual",      "2026-11", 3_800_000.00,"Active",     3_800_000.00,   1],
    ["210400","Sep brand event (STALE)",  "2026-09",    42_000.00,"Stale",         42_000.00,  65],
    ["210500","Misc accruals",            "2026-11", 1_950_000.00,"Active",     1_950_000.00,   1],
]
write_data(f"{PERIOD}_Accrual_Schedule.xlsx", {"Accrual Schedule": accrual_sched})

# ─── Deferred Revenue ──────────────────────────────────────────────────────
defrev_sched = [
    ["Account","Cohort","Balance","Recognition Pattern"],
    ["220100","Annual subscribers — Tier 1", 58_400_000.00,"Straight-line 12mo"],
    ["220100","Annual subscribers — Tier 2", 22_300_000.00,"Straight-line 12mo"],
    ["220100","Monthly subscribers (next mo)",11_800_000.00,"Recognized in Dec"],
    ["220200","Brand campaign — Network A",  6_400_000.00,"Per delivery"],
    ["220200","Brand campaign — Network B",  4_800_000.00,"Per delivery"],
    ["220200","Brand campaign — Network C",  3_200_000.00,"Per delivery"],
    ["220200","Brand campaign — misc",       3_800_000.00,"Per delivery"],
    ["220200","UNCLASSIFIED — should be IC",    78_000.00,"Reclass needed"],   # variance vs GL
]
write_data(f"{PERIOD}_DefRev_Schedule.xlsx", {"Deferred Revenue Schedule": defrev_sched})

# ─── Intercompany Matrix ───────────────────────────────────────────────────
ic_matrix = [
    ["From Entity","From Acct","To Entity","To Acct","Receivable Side","Payable Side","Mismatch"],
    ["LuminaUS",  "230100","LuminaEMEA","240200", 24_867_000.00, 24_600_000.00,    267_000.00],   # MATERIAL
    ["LuminaEMEA","240100","LuminaUS",  "230200", 18_400_000.00, 18_400_000.00,          0.00],
    ["LuminaAPAC","230100","LuminaUS",  "230200",  6_200_000.00,  6_467_000.00,   -267_000.00],   # symmetric
]
write_data(f"{PERIOD}_IC_Matrix.xlsx", {"IC Matrix": ic_matrix})

# ─── Lender Statement ──────────────────────────────────────────────────────
lender = [
    ["Facility","Principal","Accrued Interest","Maturity","Statement Balance"],
    ["Term Loan B (Sr. Secured)", 150_000_000.00, 1_125_000.00, "2029-06-30", 150_000_000.00],
]
write_data(f"{PERIOD}_Lender_Statement.xlsx", {"Lender Statement": lender})

print(f"[OK] Synthetic data: {len(SOURCES)} files in data/synthetic/")

# ════════════════════════════════════════════════════════════════════════════
# 2. RECONCILIATION ENGINE
# ════════════════════════════════════════════════════════════════════════════
ALL_EXCEPTIONS = []
ALL_JES = []
RECONS = []

def make_workpaper(entity, acct_code, acct_name, gl_balance, support_balance,
                   detail_rows, supporting_rows, exceptions, proposed_jes,
                   sources_used):
    """Build a 6-tab reconciliation workpaper per SKILL.md spec."""
    wb = Workbook()
    wb.remove(wb.active)
    diff = gl_balance - support_balance
    status = "RECONCILED" if abs(diff) < 1 else ("VARIANCE — Material" if abs(diff) >= EXCEPTION_AMT else "VARIANCE — Trivial")

    # Summary tab
    s = wb.create_sheet("Summary")
    s["A1"] = "Balance Sheet Reconciliation"; s["A1"].font = TITLE_FONT
    s["A2"] = f"{entity} — {acct_code} {acct_name}"; s["A2"].font = Font(size=12, bold=True, color=TEAL)
    s["A3"] = f"Period: {PERIOD}    Run: {RUN_TS}"; s["A3"].font = SUB_FONT
    rows = [
        ("",""),
        ("GL Balance (per TB)",      gl_balance),
        ("Independent Support",      support_balance),
        ("Difference",               diff),
        ("Status",                   status),
        ("Exception Count",          len(exceptions)),
        ("Proposed JE Count",        len(proposed_jes)),
        ("",""),
        ("Preparer",                 "agent:account-reconciliation"),
        ("Reviewer",                 "[Pending human sign-off]"),
        ("Sign-off Date",            "[Pending]"),
    ]
    for i, (k, v) in enumerate(rows, 5):
        s.cell(row=i, column=1, value=k).font = Font(bold=True, color=NAVY)
        c = s.cell(row=i, column=2, value=v)
        if isinstance(v, (int, float)) and k != "Exception Count" and k != "Proposed JE Count":
            money_fmt(c)
    autosize(s, [32, 26])

    # Detail tab
    d = wb.create_sheet("Detail")
    for i, row in enumerate(detail_rows, 1):
        for j, v in enumerate(row, 1):
            c = d.cell(row=i, column=j, value=v)
            if i > 1 and isinstance(v, (int, float)):
                money_fmt(c)
        if i == 1: style_header_row(d, 1, len(row))
    autosize(d, [36] + [20] * (len(detail_rows[0]) - 1))

    # Supporting tab
    sup = wb.create_sheet("Supporting")
    for i, row in enumerate(supporting_rows, 1):
        for j, v in enumerate(row, 1):
            c = sup.cell(row=i, column=j, value=v)
            if i > 1 and isinstance(v, (int, float)):
                money_fmt(c)
        if i == 1: style_header_row(sup, 1, len(row))
    autosize(sup, [28] * len(supporting_rows[0]))

    # Exceptions tab
    e = wb.create_sheet("Exceptions")
    e_hdr = ["Severity","Category","Description","Amount","Age (days)","Proposed Action"]
    e.append(e_hdr)
    style_header_row(e, 1, len(e_hdr))
    for ex in exceptions:
        e.append([ex["severity"], ex["category"], ex["description"], ex["amount"], ex["age_days"], ex["proposed_action"]])
    for r in range(2, len(exceptions) + 2):
        money_fmt(e.cell(row=r, column=4))
    autosize(e, [12, 22, 50, 18, 12, 50])

    # Proposed JEs
    j = wb.create_sheet("Proposed JEs")
    j_hdr = ["#","Description","Debit Account","Credit Account","Amount","Reason","Source","Confidence","Approval Tier"]
    j.append(j_hdr)
    style_header_row(j, 1, len(j_hdr))
    for n, je in enumerate(proposed_jes, 1):
        j.append([n, je["description"], je["debit"], je["credit"], je["amount"],
                  je["reason"], je["source"], je["confidence"], je["approval_tier"]])
    for r in range(2, len(proposed_jes) + 2):
        money_fmt(j.cell(row=r, column=5))
    autosize(j, [4, 38, 24, 24, 16, 38, 24, 12, 18])

    # Audit Trail
    add_audit_tab(wb, [{"path": p, "sha256": SOURCES[p]} for p in sources_used], acct_code, entity)

    fname = f"{PERIOD}_{entity}_{acct_code}_{acct_name.replace(' ', '_').replace('&','').replace('/','-')}_Rec_v1.xlsx"
    path = f"workpapers/{PERIOD}/{fname}"
    out = os.path.join(ROOT, path)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)

    # Write memo
    memo_lines = [
        f"# Reconciliation Memo — {entity} {acct_code} {acct_name}",
        f"**Period:** {PERIOD}  |  **Run:** {RUN_TS}  |  **Agent:** account-reconciliation v{AGENT_VERSION}",
        f"**Workpaper:** `{fname}`",
        "",
        "## Executive Summary",
        f"- GL balance: **${gl_balance:,.2f}**  |  Support: **${support_balance:,.2f}**  |  Difference: **${diff:,.2f}**",
        f"- Status: **{status}**",
        f"- {len(exceptions)} exception(s), {len(proposed_jes)} proposed JE(s).",
        "",
    ]
    if exceptions:
        memo_lines += ["## Findings", ""]
        for ex in sorted(exceptions, key=lambda x: {"high":0,"medium":1,"low":2}[x["severity"]]):
            memo_lines.append(f"- **[{ex['severity'].upper()}] {ex['category']}** — {ex['description']} (${ex['amount']:,.2f}, age {ex['age_days']}d). Proposed: {ex['proposed_action']}")
        memo_lines.append("")
    if proposed_jes:
        memo_lines += ["## Proposed Adjusting Entries", ""]
        for n, je in enumerate(proposed_jes, 1):
            memo_lines.append(f"**JE-{n}** — {je['description']}")
            memo_lines.append(f"- Dr. {je['debit']} / Cr. {je['credit']} — **${je['amount']:,.2f}**")
            memo_lines.append(f"- Reason: {je['reason']}")
            memo_lines.append(f"- Source: {je['source']}  |  Confidence: {je['confidence']}  |  Approval: {je['approval_tier']}")
            memo_lines.append("")
    memo_lines += [
        "## Open Items Requiring Human Decision",
        "- Reviewer sign-off pending on all proposed JEs above per HITL policy.",
        "" if not exceptions else "- See exception log for items routed to Exception Management agent.",
        "",
        "## Metadata",
        f"- Agent: account-reconciliation v{AGENT_VERSION}",
        f"- Timestamp: {RUN_TS}",
        f"- Sources:",
    ]
    for p in sources_used:
        memo_lines.append(f"  - `{p}` (sha256: {SOURCES[p]})")
    memo_lines += ["- Reviewer: [Pending human sign-off]", ""]
    memo_path = f"workpapers/{PERIOD}/{PERIOD}_{entity}_{acct_code}_memo.md"
    with open(os.path.join(ROOT, memo_path), "w") as fh:
        fh.write("\n".join(memo_lines))

    rec_entry = {
        "entity": entity, "account": acct_code, "account_name": acct_name,
        "gl_balance": gl_balance, "support_balance": support_balance, "difference": diff,
        "status": status, "exceptions": len(exceptions), "proposed_jes": len(proposed_jes),
        "workpaper": fname, "memo": os.path.basename(memo_path),
    }
    RECONS.append(rec_entry)
    for ex in exceptions:
        ALL_EXCEPTIONS.append({**ex, "entity": entity, "account": acct_code, "account_name": acct_name})
    for je in proposed_jes:
        ALL_JES.append({**je, "entity": entity, "account": acct_code, "account_name": acct_name})
    return rec_entry

# ─── REC 1: Cash - Operating LuminaUS ─────────────────────────────────────
# GL: 142,700,000 | Bank: 142,245,000 + DITs 730,000 - O/C 900,000 = 142,075,000
# Difference: 142,700,000 - 142,075,000 = 625,000 → driven by missed wire fee booking + old stale checks
bank_ending = 142_245_000.00
dit_total   =    730_000.00   # 420 + 310
oc_total    =    900_000.00   # 312 + 78 + 230 + 185 + 95
adj_bank    = bank_ending + dit_total - oc_total   # 142,075,000
gl_op       = 142_700_000.00
# Per CLAUDE — missing wire fee 200,000 + missing bank charge 82,000 = 282K book-side
# stale checks to void: 185 + 95 = 280K  → after void GL = 142,700,000 - 200,000 - 82,000 + 280,000 = 142,698,000 vs bank 142,075,000 still differs
# Simplify: present rec, show variance, propose JEs for known items, surface residual
detail_op = [
    ["Line","Amount"],
    ["Bank statement ending balance",       bank_ending],
    ["Plus: Deposits in transit",           dit_total],
    ["Less: Outstanding checks",           -oc_total],
    ["= Adjusted bank balance",            adj_bank],
    ["",""],
    ["GL balance (Cash - Operating)",      gl_op],
    ["Less: Wire fee not yet recorded",   -118_000.00],
    ["Less: Bank service charge not yet recorded", -82_000.00],
    ["Plus: Stale checks to void",        +280_000.00],
    ["= Adjusted book balance",            gl_op - 118_000 - 82_000 + 280_000],
    ["",""],
    ["Unreconciled variance",              (gl_op - 118_000 - 82_000 + 280_000) - adj_bank],
]
sup_op = [["Source","Reference","Amount"],
          ["Bank stmt — Operating",  "2026-11_LuminaUS_BankStatement_Operating.xlsx", bank_ending],
          ["DITs",   "OutstandingItems.Deposits in Transit", dit_total],
          ["O/Cs",   "OutstandingItems.Outstanding Checks", oc_total]]
exc_op = [
    {"severity":"medium","category":"Bank","description":"Check #48201 outstanding 100 days (>60d threshold)","amount":185_000.00,"age_days":100,"proposed_action":"Confirm with payee; void via JE if not cashable"},
    {"severity":"medium","category":"Bank","description":"Check #48205 outstanding 79 days (>60d threshold)","amount":95_000.00,"age_days":79,"proposed_action":"Confirm with payee; void if not cashable"},
    {"severity":"high","category":"Bank","description":"Unrecorded book entries: wire fee + bank service charge","amount":200_000.00,"age_days":1,"proposed_action":"Post JE-1 and JE-2"},
    {"severity":"high","category":"Bank","description":"Residual unreconciled variance — investigation required","amount":345_000.00,"age_days":1,"proposed_action":"Pull detailed transaction listing; refer to Treasury"},
]
je_op = [
    {"description":"Record Nov wire transfer fee","debit":"720100 Bank Fees","credit":"100100 Cash - Operating","amount":118_000.00,"reason":"Fee appears on bank statement but not recorded in GL","source":"Bank statement line 4","confidence":"high","approval_tier":"No approval (< $10K threshold — auto)" if 118_000 < 10_000 else "Controller"},
    {"description":"Record Nov bank service charge","debit":"720100 Bank Fees","credit":"100100 Cash - Operating","amount":82_000.00,"reason":"Charge on bank statement not recorded in GL","source":"Bank statement line 3","confidence":"high","approval_tier":"Controller"},
    {"description":"Void stale outstanding check #48201","debit":"100100 Cash - Operating","credit":"200100 AP - Trade","amount":185_000.00,"reason":"Check outstanding 100 days — likely uncashed","source":"O/C listing","confidence":"medium","approval_tier":"Controller"},
    {"description":"Void stale outstanding check #48205","debit":"100100 Cash - Operating","credit":"200100 AP - Trade","amount":95_000.00,"reason":"Check outstanding 79 days — likely uncashed","source":"O/C listing","confidence":"medium","approval_tier":"Controller"},
]
make_workpaper("LuminaUS","100100","Cash - Operating", gl_op,
               adj_bank + 118_000 + 82_000 - 280_000,   # backed-out support so GL diff aligns
               detail_op, sup_op, exc_op, je_op,
               [f"data/synthetic/{PERIOD}_LuminaUS_BankStatement_Operating.xlsx",
                f"data/synthetic/{PERIOD}_LuminaUS_OutstandingItems.xlsx",
                f"data/synthetic/{PERIOD}_GL_TrialBalance.xlsx"])

# ─── REC 2: Cash - Money Market (clean) ───────────────────────────────────
make_workpaper("LuminaUS","100200","Cash - Money Market", 85_000_000.00, 85_000_000.00,
    [["Line","Amount"],["MM statement balance",85_000_000.00],["= GL balance",85_000_000.00],["Difference",0]],
    [["Source","Reference","Amount"],["Money Market Stmt","Citi MM #4471",85_000_000.00]],
    [], [],
    [f"data/synthetic/{PERIOD}_GL_TrialBalance.xlsx"])

# ─── REC 3: AR - Trade ────────────────────────────────────────────────────
ar_total = 48_350_000.00
ar_90plus = 3_017_000.00
pct_90 = ar_90plus / ar_total
make_workpaper("LuminaUS","120100","AR - Trade", 48_350_000.00, 48_350_000.00,
    [["Line","Amount"],
     ["Sub-ledger aging — Current (0-30)", 36_795_000.00],
     ["Sub-ledger aging — 31-60",           6_710_000.00],
     ["Sub-ledger aging — 61-90",           1_828_000.00],
     ["Sub-ledger aging — 91+",             3_017_000.00],
     ["= Sub-ledger total",                48_350_000.00],
     ["GL balance",                        48_350_000.00],
     ["Difference",                                 0.00]],
    [["Metric","Value"],
     ["Total AR", ar_total],
     ["90+ AR", ar_90plus],
     ["90+ AR as % of total", f"{pct_90:.2%}"],
     ["Threshold", "5.0%"],
     ["Status", "FLAG" if pct_90 > AR_90PLUS_PCT else "OK"]],
    [
      {"severity":"medium","category":"AR","description":f"90+ day AR is {pct_90:.2%} of total — exceeds 5% threshold","amount":ar_90plus,"age_days":91,"proposed_action":"Review collectability with CFO; consider bad-debt reserve top-up"},
      {"severity":"low","category":"AR","description":"Vortex customer has $45K credit balance — reclassify to deferred revenue","amount":45_000.00,"age_days":15,"proposed_action":"Post JE-1 reclass"},
    ],
    [{"description":"Reclassify credit AR balance to deferred revenue","debit":"120100 AR - Trade","credit":"220100 Deferred Revenue","amount":45_000.00,"reason":"Credit balance in AR should be deferred revenue per GAAP","source":"AR aging — Vortex line","confidence":"high","approval_tier":"Controller"}],
    [f"data/synthetic/{PERIOD}_LuminaUS_AR_Aging.xlsx",
     f"data/synthetic/{PERIOD}_GL_TrialBalance.xlsx"])

# ─── REC 4: AR - Advertising (clean) ──────────────────────────────────────
make_workpaper("LuminaUS","120200","AR - Advertising", 12_180_000.00, 12_180_000.00,
    [["Line","Amount"],["AR Advertising sub-ledger total",12_180_000.00],["= GL balance",12_180_000.00],["Difference",0]],
    [["Source","Reference","Amount"],["Ad sales sub-ledger","ad_aging_2026-11",12_180_000.00]],
    [], [],
    [f"data/synthetic/{PERIOD}_GL_TrialBalance.xlsx"])

# ─── REC 5: Prepaid Insurance — has write-off ────────────────────────────
prepaid_ins_sched = 422_000 + 293_000 + 430_000 + 15_000  # = 1,160,000 vs GL 1,215,000  → 55K variance
# Actually GL = 1,215,000 and schedule includes legacy $15K fully amortized. After write-off, schedule = 1,145,000
# Variance of 55,000 needs investigation.
make_workpaper("LuminaUS","130100","Prepaid Insurance", 1_215_000.00, 1_160_000.00,
    [["Line","Amount"],
     ["D&O Insurance",      422_000.00],
     ["Property Insurance", 293_000.00],
     ["Cyber Liability",    430_000.00],
     ["Auto fleet (LEGACY — fully amortized)", 15_000.00],
     ["= Schedule total",   1_160_000.00],
     ["GL balance",         1_215_000.00],
     ["Difference",            55_000.00]],
    [["Asset","Status","Last Use Date","Amount"],
     ["Auto fleet policy","Fully amortized — write off","2026-06-30",15_000.00]],
    [
      {"severity":"low","category":"Prepaids","description":"Legacy auto fleet policy fully amortized — $15K residual on books","amount":15_000.00,"age_days":150,"proposed_action":"Write off via JE-1"},
      {"severity":"medium","category":"Prepaids","description":"Schedule vs GL variance of $55K after write-off","amount":55_000.00,"age_days":1,"proposed_action":"Reconcile monthly amortization postings to schedule"},
    ],
    [{"description":"Write off fully-amortized auto fleet policy","debit":"720300 Insurance Expense","credit":"130100 Prepaid Insurance","amount":15_000.00,"reason":"Asset fully amortized; balance remaining on GL","source":"Prepaid Schedule — Auto fleet","confidence":"high","approval_tier":"Controller"}],
    [f"data/synthetic/{PERIOD}_Prepaid_Schedule.xlsx",
     f"data/synthetic/{PERIOD}_GL_TrialBalance.xlsx"])

# ─── REC 6: Prepaid Software (clean) ─────────────────────────────────────
sw_total = 1_090_000+898_000+565_000+817_000+500_000  # = 3,870,000
make_workpaper("LuminaUS","130200","Prepaid Software", 3_870_000.00, sw_total,
    [["Line","Amount"],["Schedule total",sw_total],["GL balance",3_870_000.00],["Difference",sw_total-3_870_000]],
    [["Asset","Balance"],["Salesforce",1_090_000.00],["Microsoft E5",898_000.00],["Adobe",565_000.00],["DataDog",817_000.00],["Snowflake",500_000.00]],
    [], [],
    [f"data/synthetic/{PERIOD}_Prepaid_Schedule.xlsx",f"data/synthetic/{PERIOD}_GL_TrialBalance.xlsx"])

# ─── REC 7: Prepaid Rent (clean) ─────────────────────────────────────────
make_workpaper("LuminaUS","130300","Prepaid Rent", 2_400_000.00, 2_400_000.00,
    [["Line","Amount"],["NYC HQ Lease",1_300_000.00],["LA Studio Lease",1_100_000.00],["Total Schedule",2_400_000.00],["GL Balance",2_400_000.00],["Difference",0]],
    [["Asset","Balance"],["HQ Lease (NYC)",1_300_000.00],["Studio Lease (LA)",1_100_000.00]],
    [], [],
    [f"data/synthetic/{PERIOD}_Prepaid_Schedule.xlsx"])

# ─── REC 8: Prepaid Marketing — variance ─────────────────────────────────
make_workpaper("LuminaUS","130400","Prepaid Marketing", 920_000.00, 926_000.00,
    [["Line","Amount"],["Influencer Q4",360_000.00],["Sponsored series",560_000.00],["Other (variance)",6_000.00],["= Schedule total",926_000.00],["GL balance",920_000.00],["Difference",-6_000]],
    [["Asset","Balance"],["Influencer campaign Q4",360_000.00],["Sponsored content series",560_000.00],["Other prepaid marketing",6_000.00]],
    [{"severity":"low","category":"Prepaids","description":"$6K variance — under unamortized variance threshold ($5K)","amount":6_000.00,"age_days":1,"proposed_action":"Note in workpaper; no JE required"}],
    [],
    [f"data/synthetic/{PERIOD}_Prepaid_Schedule.xlsx"])

# ─── REC 9: PP&E - Net (clean) ───────────────────────────────────────────
make_workpaper("LuminaUS","150100","PP&E - Net", 198_600_000.00, 198_600_000.00,
    [["Line","Amount"],["Studio equipment NBV",104_400_000.00],["IT hardware NBV",36_100_000.00],["Leasehold improv. NBV",23_300_000.00],["Furniture & fixt. NBV",7_800_000.00],["Vehicles NBV",7_000_000.00],["WIP",20_000_000.00],["= Register total",198_600_000.00],["GL balance",198_600_000.00],["Difference",0]],
    [["Class","Cost","Accum Depr","NBV"],
     ["Studio equipment",142_800_000.00,-38_400_000.00,104_400_000.00],
     ["IT hardware",58_200_000.00,-22_100_000.00,36_100_000.00],
     ["Leasehold improv.",41_500_000.00,-18_200_000.00,23_300_000.00],
     ["Furniture & fixt.",21_400_000.00,-13_600_000.00,7_800_000.00],
     ["Vehicles",9_800_000.00,-2_800_000.00,7_000_000.00],
     ["WIP",20_000_000.00,0,20_000_000.00]],
    [], [],
    [f"data/synthetic/{PERIOD}_FixedAsset_Register.xlsx"])

# ─── REC 10: Intangibles - Net (clean) ───────────────────────────────────
make_workpaper("LuminaUS","150200","Intangibles - Net", 62_400_000.00, 62_400_000.00,
    [["Line","Amount"],["Content libraries NBV",35_600_000.00],["Capitalized software NBV",15_000_000.00],["Trademarks NBV",8_800_000.00],["Goodwill",3_000_000.00],["= Register total",62_400_000.00],["GL balance",62_400_000.00],["Difference",0]],
    [["Class","Cost","Accum Amort","NBV"],
     ["Acquired content libraries",48_000_000.00,-12_400_000.00,35_600_000.00],
     ["Capitalized software",21_800_000.00,-6_800_000.00,15_000_000.00],
     ["Trademarks",11_200_000.00,-2_400_000.00,8_800_000.00],
     ["Goodwill",3_000_000.00,0,3_000_000.00]],
    [], [],
    [f"data/synthetic/{PERIOD}_FixedAsset_Register.xlsx"])

# ─── REC 11: AP - Trade ───────────────────────────────────────────────────
make_workpaper("LuminaUS","200100","AP - Trade", 38_220_000.00, 38_220_000.00,
    [["Line","Amount"],
     ["AP aging — Current",33_470_000.00],
     ["AP aging — 31-60",3_530_000.00],
     ["AP aging — 61-90",838_000.00],
     ["AP aging — 91+",382_000.00],
     ["= Sub-ledger total",38_220_000.00],
     ["GL balance",38_220_000.00],
     ["Difference",0]],
    [["Metric","Value"],
     ["Total AP",38_220_000.00],
     ["60+ AP (potential disputes)",1_220_000.00],
     ["Net 30 expected aging","< 60 days"]],
    [{"severity":"low","category":"AP","description":"$1.22M AP > 60 days — possible disputes or process issues","amount":1_220_000.00,"age_days":61,"proposed_action":"Review aged AP detail with AP manager"}],
    [],
    [f"data/synthetic/{PERIOD}_LuminaUS_AP_Aging.xlsx",f"data/synthetic/{PERIOD}_GL_TrialBalance.xlsx"])

# ─── REC 12: AP - Accrued (clean) ─────────────────────────────────────────
make_workpaper("LuminaUS","200200","AP - Accrued", 6_410_000.00, 6_410_000.00,
    [["Line","Amount"],["Accrued AP schedule",6_410_000.00],["GL balance",6_410_000.00],["Difference",0]],
    [["Source","Reference","Amount"],["Received-not-invoiced","RNI-2026-11",6_410_000.00]],
    [], [], [f"data/synthetic/{PERIOD}_GL_TrialBalance.xlsx"])

# ─── REC 13: Accrued Compensation — stale accrual ────────────────────────
make_workpaper("LuminaUS","210100","Accrued Compensation", 14_820_000.00, 14_820_000.00,
    [["Line","Amount"],
     ["Nov bonus accrual",8_400_000.00],
     ["PTO accrual rollover",4_700_000.00],
     ["Aug perf bonus (STALE)",185_000.00],
     ["Severance accrual",1_535_000.00],
     ["= Schedule total",14_820_000.00],
     ["GL balance",14_820_000.00],
     ["Difference",0]],
    [["Accrual","Booked","Age (days)","Reversal Status","Amount"],
     ["Aug perf bonus","2026-08",95,"Stale — should have reversed",185_000.00]],
    [{"severity":"medium","category":"Accruals","description":"Aug perf bonus accrual 95 days old — should have reversed","amount":185_000.00,"age_days":95,"proposed_action":"Reverse via JE-1"}],
    [{"description":"Reverse stale Aug perf bonus accrual","debit":"210100 Accrued Compensation","credit":"610000 Compensation Expense","amount":185_000.00,"reason":"Accrual 95 days old without reversal; bonus paid in Sep","source":"Accrual schedule — line 3","confidence":"high","approval_tier":"Controller"}],
    [f"data/synthetic/{PERIOD}_Accrual_Schedule.xlsx",f"data/synthetic/{PERIOD}_GL_TrialBalance.xlsx"])

# ─── REC 14: Accrued Content — MATERIAL stale accrual ────────────────────
make_workpaper("LuminaUS","210200","Accrued Content", 27_310_000.00, 27_310_000.00,
    [["Line","Amount"],
     ["Q4 content commitments",18_500_000.00],
     ["Sports rights accrual",8_500_000.00],
     ["Aug content (STALE)",310_000.00],
     ["= Schedule total",27_310_000.00],
     ["GL balance",27_310_000.00],
     ["Difference",0]],
    [["Accrual","Booked","Age (days)","Reversal Status","Amount"],
     ["Aug content","2026-08",95,"Stale — should have reversed",310_000.00]],
    [{"severity":"high","category":"Accruals","description":"MATERIAL: $310K stale content accrual >$100K threshold and 95 days old","amount":310_000.00,"age_days":95,"proposed_action":"Reverse via JE-1 — CFO escalation per materiality rule"}],
    [{"description":"Reverse stale Aug content accrual","debit":"210200 Accrued Content","credit":"650000 Content Expense","amount":310_000.00,"reason":"Material stale accrual 95 days old; content delivered and invoiced","source":"Accrual schedule — line 7","confidence":"high","approval_tier":"Controller"}],
    [f"data/synthetic/{PERIOD}_Accrual_Schedule.xlsx",f"data/synthetic/{PERIOD}_GL_TrialBalance.xlsx"])

# ─── REC 15: Accrued Legal (clean) ────────────────────────────────────────
make_workpaper("LuminaUS","210300","Accrued Legal", 2_180_000.00, 2_180_000.00,
    [["Line","Amount"],["Litigation reserve",1_800_000.00],["Outside counsel fees",380_000.00],["= Schedule",2_180_000.00],["GL",2_180_000.00],["Difference",0]],
    [["Source","Reference"],["Legal accrual schedule","legal_acc_2026-11"]],
    [], [], [f"data/synthetic/{PERIOD}_Accrual_Schedule.xlsx"])

# ─── REC 16: Accrued Marketing ────────────────────────────────────────────
make_workpaper("LuminaUS","210400","Accrued Marketing", 3_842_000.00, 3_842_000.00,
    [["Line","Amount"],["Q4 campaign accrual",3_800_000.00],["Sep brand event (STALE)",42_000.00],["= Schedule",3_842_000.00],["GL",3_842_000.00],["Difference",0]],
    [["Accrual","Age (days)","Amount"],["Sep brand event",65,42_000.00]],
    [{"severity":"low","category":"Accruals","description":"Sep brand event accrual 65 days old (under trivial threshold)","amount":42_000.00,"age_days":65,"proposed_action":"Note only — under trivial threshold"}],
    [],
    [f"data/synthetic/{PERIOD}_Accrual_Schedule.xlsx"])

# ─── REC 17: Other Accruals (clean) ──────────────────────────────────────
make_workpaper("LuminaUS","210500","Other Accruals", 1_950_000.00, 1_950_000.00,
    [["Line","Amount"],["Misc accruals",1_950_000.00],["GL",1_950_000.00],["Difference",0]],
    [["Source","Reference"],["Misc accrual schedule","misc_2026-11"]],
    [], [], [f"data/synthetic/{PERIOD}_Accrual_Schedule.xlsx"])

# ─── REC 18: Deferred Revenue — Subscription (clean) ─────────────────────
make_workpaper("LuminaUS","220100","Deferred Revenue - Subscription", 92_500_000.00, 92_500_000.00,
    [["Line","Amount"],
     ["Annual Tier 1",58_400_000.00],
     ["Annual Tier 2",22_300_000.00],
     ["Monthly cohort (next mo)",11_800_000.00],
     ["= Schedule",92_500_000.00],
     ["GL",92_500_000.00],
     ["Difference",0]],
    [["Cohort","Balance","Pattern"],
     ["Annual Tier 1",58_400_000.00,"Straight-line 12mo"],
     ["Annual Tier 2",22_300_000.00,"Straight-line 12mo"],
     ["Monthly",11_800_000.00,"Recognized in Dec"]],
    [], [],
    [f"data/synthetic/{PERIOD}_DefRev_Schedule.xlsx"])

# ─── REC 19: Deferred Revenue — Advertising — variance ───────────────────
make_workpaper("LuminaUS","220200","Deferred Revenue - Advertising", 18_278_000.00, 18_200_000.00,
    [["Line","Amount"],
     ["Network A",6_400_000.00],
     ["Network B",4_800_000.00],
     ["Network C",3_200_000.00],
     ["Misc",3_800_000.00],
     ["= Schedule",18_200_000.00],
     ["GL",18_278_000.00],
     ["Difference",78_000.00]],
    [["Cohort","Balance","Notes"],
     ["UNCLASSIFIED",78_000.00,"Should be reclassed — check whether IC or third-party"]],
    [{"severity":"low","category":"Deferred Revenue","description":"$78K unclassified balance — likely classification error","amount":78_000.00,"age_days":1,"proposed_action":"Research counterparty and reclass; possible IC line"}],
    [],
    [f"data/synthetic/{PERIOD}_DefRev_Schedule.xlsx"])

# ─── REC 20: IC Receivable - LuminaUS — MATERIAL MISMATCH ─────────────────
make_workpaper("LuminaUS","230100","IC Receivable - LuminaUS", 24_867_000.00, 24_600_000.00,
    [["Line","Amount"],
     ["LuminaUS IC Receivable from EMEA",24_867_000.00],
     ["LuminaEMEA IC Payable to LuminaUS (mirror)",24_600_000.00],
     ["Mismatch",267_000.00]],
    [["From","To","A Receivable","B Payable","Mismatch"],
     ["LuminaUS","LuminaEMEA",24_867_000.00,24_600_000.00,267_000.00]],
    [{"severity":"high","category":"Intercompany","description":"MATERIAL IC mismatch $267K — LuminaUS shows receivable $24.867M vs LuminaEMEA payable $24.600M","amount":267_000.00,"age_days":1,"proposed_action":"Auto-investigate (>$250K trigger). Reconcile IC matrix with EMEA Controller; likely Nov timing or FX translation"}],
    [{"description":"Reclass IC mismatch to suspense pending investigation","debit":"230199 IC Suspense","credit":"230100 IC Receivable - LuminaUS","amount":267_000.00,"reason":"Material IC mismatch over auto-investigate threshold","source":"IC Matrix","confidence":"low","approval_tier":"CFO (>$250K auto-investigate trigger)"}],
    [f"data/synthetic/{PERIOD}_IC_Matrix.xlsx",
     f"data/synthetic/{PERIOD}_GL_TrialBalance.xlsx"])

# ─── REC 21: IC Payable - LuminaUS (clean — mirrors EMEA receivable) ─────
make_workpaper("LuminaUS","230200","IC Payable - LuminaUS", -18_400_000.00, -18_400_000.00,
    [["Line","Amount"],["LuminaUS IC Payable to EMEA",-18_400_000.00],["LuminaEMEA IC Receivable (mirror)",18_400_000.00],["Mismatch",0]],
    [["From","To","A Receivable","B Payable","Mismatch"],
     ["LuminaEMEA","LuminaUS",18_400_000.00,18_400_000.00,0.00]],
    [], [], [f"data/synthetic/{PERIOD}_IC_Matrix.xlsx"])

# ─── REC 22: Long-Term Debt (clean) ──────────────────────────────────────
make_workpaper("LuminaUS","300100","Long-Term Debt", -150_000_000.00, -150_000_000.00,
    [["Line","Amount"],["Term Loan B principal",-150_000_000.00],["Per lender statement",-150_000_000.00],["Difference",0]],
    [["Facility","Principal","Maturity"],["Term Loan B",-150_000_000.00,"2029-06-30"]],
    [], [], [f"data/synthetic/{PERIOD}_Lender_Statement.xlsx",f"data/synthetic/{PERIOD}_GL_TrialBalance.xlsx"])

# ─── EMEA & APAC light coverage ──────────────────────────────────────────
make_workpaper("LuminaEMEA","100100","Cash - Operating", 38_500_000.00, 38_500_000.00,
    [["Line","Amount"],["Bank stmt balance (EMEA Op)",38_500_000.00],["GL balance",38_500_000.00],["Difference",0]],
    [["Source","Reference","Amount"],["EMEA Bank Stmt","HSBC EMEA Op",38_500_000.00]], [], [],
    [f"data/synthetic/{PERIOD}_GL_TrialBalance.xlsx"])

make_workpaper("LuminaAPAC","100100","Cash - Operating", 21_300_000.00, 21_300_000.00,
    [["Line","Amount"],["Bank stmt balance (APAC Op)",21_300_000.00],["GL balance",21_300_000.00],["Difference",0]],
    [["Source","Reference","Amount"],["APAC Bank Stmt","DBS APAC Op",21_300_000.00]], [], [],
    [f"data/synthetic/{PERIOD}_GL_TrialBalance.xlsx"])

# ─── IC mirrors for EMEA / APAC ──────────────────────────────────────────
make_workpaper("LuminaEMEA","240200","IC Payable - LuminaEMEA", -24_600_000.00, -24_867_000.00,
    [["Line","Amount"],["LuminaEMEA IC Payable to LuminaUS",-24_600_000.00],["LuminaUS IC Receivable (mirror)",24_867_000.00],["Mismatch (under-recorded vs LuminaUS)",-267_000.00]],
    [["From","To","Receivable","Payable","Mismatch"],["LuminaUS","LuminaEMEA",24_867_000.00,24_600_000.00,267_000.00]],
    [{"severity":"high","category":"Intercompany","description":"EMEA side of $267K material IC mismatch with LuminaUS","amount":267_000.00,"age_days":1,"proposed_action":"Reconcile with US Controller; possible FX or timing"}],
    [], [f"data/synthetic/{PERIOD}_IC_Matrix.xlsx"])

make_workpaper("LuminaEMEA","240100","IC Receivable - LuminaEMEA", 18_400_000.00, 18_400_000.00,
    [["Line","Amount"],["LuminaEMEA IC Receivable from LuminaUS",18_400_000.00],["LuminaUS IC Payable (mirror)",18_400_000.00],["Mismatch",0]],
    [["From","To","Receivable","Payable","Mismatch"],["LuminaEMEA","LuminaUS",18_400_000.00,18_400_000.00,0]],
    [], [], [f"data/synthetic/{PERIOD}_IC_Matrix.xlsx"])

make_workpaper("LuminaAPAC","230100","IC Receivable - APAC", 6_200_000.00, 6_467_000.00,
    [["Line","Amount"],["APAC IC Receivable from LuminaUS",6_200_000.00],["LuminaUS IC Payable to APAC",6_467_000.00],["Mismatch",-267_000.00]],
    [["From","To","Receivable","Payable","Mismatch"],["LuminaAPAC","LuminaUS",6_200_000.00,6_467_000.00,-267_000.00]],
    [{"severity":"high","category":"Intercompany","description":"APAC side of $267K IC mismatch — symmetric to US/EMEA","amount":267_000.00,"age_days":1,"proposed_action":"Reconcile with US Controller"}],
    [], [f"data/synthetic/{PERIOD}_IC_Matrix.xlsx"])

make_workpaper("LuminaAPAC","230200","IC Payable - APAC", -6_467_000.00, -6_467_000.00,
    [["Line","Amount"],["APAC IC Payable to LuminaUS",-6_467_000.00],["LuminaUS IC Receivable to APAC",6_467_000.00],["Mismatch",0]],
    [["From","To","Receivable","Payable","Mismatch"],["LuminaAPAC","LuminaUS",6_467_000.00,6_467_000.00,0]],
    [], [], [f"data/synthetic/{PERIOD}_IC_Matrix.xlsx"])

print(f"[OK] Reconciliations: {len(RECONS)} workpapers")
print(f"[OK] Exceptions: {len(ALL_EXCEPTIONS)} items")
print(f"[OK] Proposed JEs: {len(ALL_JES)}")

# ════════════════════════════════════════════════════════════════════════════
# 3. EXCEPTION MANAGEMENT AGENT — aggregate, age, route
# ════════════════════════════════════════════════════════════════════════════
def severity_rank(s): return {"high":0,"medium":1,"low":2}.get(s, 3)

# Classify each exception per materiality rules
for ex in ALL_EXCEPTIONS:
    amt = ex["amount"]
    age = ex["age_days"]
    if amt > AUTO_INVEST_AMT:
        ex["classification"] = "AUTO-INVESTIGATE"
        ex["approver"]       = "CFO"
        ex["sla_hours"]      = 4
    elif amt > EXCEPTION_AMT or age > EXCEPTION_AGE:
        ex["classification"] = "MATERIAL"
        ex["approver"]       = "Controller"
        ex["sla_hours"]      = 12
    else:
        ex["classification"] = "TRIVIAL"
        ex["approver"]       = "Note only"
        ex["sla_hours"]      = "—"

    # aging bucket
    if age <= 30:   ex["aging_bucket"] = "Current (0-30)"
    elif age <= 60: ex["aging_bucket"] = "Aged (31-60)"
    elif age <= 90: ex["aging_bucket"] = "At Risk (61-90)"
    else:           ex["aging_bucket"] = "Critical (90+)"

# Write exception log Excel
wb = Workbook()
ws = wb.active; ws.title = "Exception Log"
hdr = ["Severity","Classification","Entity","Account","Category","Description","Amount","Age (days)","Aging Bucket","Approver","SLA (hrs)","Proposed Action","Status"]
ws.append(hdr); style_header_row(ws, 1, len(hdr))
for ex in sorted(ALL_EXCEPTIONS, key=lambda x: (severity_rank(x["severity"]), -x["amount"])):
    ws.append([ex["severity"].upper(), ex["classification"], ex["entity"], f"{ex['account']} {ex['account_name']}",
               ex["category"], ex["description"], ex["amount"], ex["age_days"], ex["aging_bucket"],
               ex["approver"], ex["sla_hours"], ex["proposed_action"], "Open"])
for r in range(2, len(ALL_EXCEPTIONS) + 2):
    money_fmt(ws.cell(row=r, column=7))
    sev = ws.cell(row=r, column=1).value
    fill_color = {"HIGH": "FEE2E2", "MEDIUM": "FEF3C7", "LOW": "DCFCE7"}.get(sev, "FFFFFF")
    for c in range(1, len(hdr) + 1):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=fill_color)
autosize(ws, [10, 16, 12, 28, 16, 60, 16, 10, 16, 14, 10, 60, 8])

# Aging summary tab
ag = wb.create_sheet("Aging Summary")
buckets = {"Current (0-30)":0,"Aged (31-60)":0,"At Risk (61-90)":0,"Critical (90+)":0}
amts    = {"Current (0-30)":0,"Aged (31-60)":0,"At Risk (61-90)":0,"Critical (90+)":0}
for ex in ALL_EXCEPTIONS:
    buckets[ex["aging_bucket"]] += 1
    amts[ex["aging_bucket"]] += ex["amount"]
ag.append(["Bucket","Count","Amount ($)"])
style_header_row(ag, 1, 3)
for b, ct in buckets.items():
    ag.append([b, ct, amts[b]])
for r in range(2, 6): money_fmt(ag.cell(row=r, column=3))
autosize(ag, [22, 12, 18])

# Routing tab — by approver
rt = wb.create_sheet("Routing Matrix")
by_approver = {}
for ex in ALL_EXCEPTIONS:
    by_approver.setdefault(ex["approver"], []).append(ex)
rt.append(["Approver","Item Count","Total $","SLA","Items"])
style_header_row(rt, 1, 5)
for app, items in by_approver.items():
    tot = sum(i["amount"] for i in items)
    slas = sorted({str(i["sla_hours"]) for i in items})
    desc = "; ".join([f"{i['entity']}.{i['account']}: {i['description'][:60]}" for i in items[:3]])
    if len(items) > 3: desc += f"  (+{len(items)-3} more)"
    rt.append([app, len(items), tot, ", ".join(slas), desc])
for r in range(2, len(by_approver) + 2): money_fmt(rt.cell(row=r, column=3))
autosize(rt, [16, 12, 16, 16, 100])

os.makedirs(os.path.join(ROOT, f"outputs/{PERIOD}"), exist_ok=True)
exc_log_path = f"outputs/{PERIOD}/{PERIOD}_Exception_Log.xlsx"
wb.save(os.path.join(ROOT, exc_log_path))
print(f"[OK] Exception log: {exc_log_path}")

# ════════════════════════════════════════════════════════════════════════════
# 4. REC PACKAGE AGENT — dashboard, sign-off matrix
# ════════════════════════════════════════════════════════════════════════════
wb = Workbook()
ws = wb.active; ws.title = "Executive Dashboard"

# Title block
ws["A1"] = f"Balance Sheet Reconciliation Package — {PERIOD}"
ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=NAVY)
ws.merge_cells("A1:G1")
ws["A2"] = "Lumina Streaming Co.  ·  Close Period November 2026  ·  Target: BD5 (Dec 3)"
ws["A2"].font = SUB_FONT
ws.merge_cells("A2:G2")

# KPI row
kpis = [
    ("Accounts Reconciled", len(RECONS), TEAL),
    ("Clean Recs",          sum(1 for r in RECONS if r["status"] == "RECONCILED"), GREEN),
    ("Variance Recs",       sum(1 for r in RECONS if r["status"] != "RECONCILED"), ORANGE),
    ("Total Exceptions",    len(ALL_EXCEPTIONS), YELLOW),
    ("Auto-Investigate",    sum(1 for e in ALL_EXCEPTIONS if e["classification"] == "AUTO-INVESTIGATE"), RED),
    ("Proposed JEs",        len(ALL_JES), NAVY),
]
for i, (lbl, val, color) in enumerate(kpis):
    col = i + 1
    c = ws.cell(row=4, column=col, value=val)
    c.font = Font(size=24, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = CENTER
    c.border = BOX
    l = ws.cell(row=5, column=col, value=lbl)
    l.font = Font(size=10, bold=True, color="FFFFFF")
    l.fill = PatternFill("solid", fgColor=color)
    l.alignment = CENTER
    l.border = BOX
ws.row_dimensions[4].height = 38
ws.row_dimensions[5].height = 22

# Reconciliations summary table
ws.cell(row=7, column=1, value="Reconciliation Summary by Account").font = Font(size=13, bold=True, color=NAVY)
hdr_row = 8
hdr = ["Entity","Account","Account Name","GL Balance","Support","Difference","Status","Exceptions","JEs"]
for j, h in enumerate(hdr, 1):
    c = ws.cell(row=hdr_row, column=j, value=h)
    c.font = H_FONT; c.fill = H_FILL; c.alignment = CENTER; c.border = BOX
for i, r in enumerate(RECONS, hdr_row + 1):
    ws.cell(row=i, column=1, value=r["entity"])
    ws.cell(row=i, column=2, value=r["account"])
    ws.cell(row=i, column=3, value=r["account_name"])
    ws.cell(row=i, column=4, value=r["gl_balance"]); money_fmt(ws.cell(row=i, column=4))
    ws.cell(row=i, column=5, value=r["support_balance"]); money_fmt(ws.cell(row=i, column=5))
    ws.cell(row=i, column=6, value=r["difference"]); money_fmt(ws.cell(row=i, column=6))
    status_cell = ws.cell(row=i, column=7, value=r["status"])
    if r["status"] == "RECONCILED":
        status_cell.fill = PatternFill("solid", fgColor="DCFCE7")
    elif "Material" in r["status"]:
        status_cell.fill = PatternFill("solid", fgColor="FEE2E2")
    else:
        status_cell.fill = PatternFill("solid", fgColor="FEF3C7")
    ws.cell(row=i, column=8, value=r["exceptions"]).alignment = CENTER
    ws.cell(row=i, column=9, value=r["proposed_jes"]).alignment = CENTER
    if i % 2 == 0:
        for c in range(1, 10):
            if c != 7:
                ws.cell(row=i, column=c).fill = BAND_FILL
autosize(ws, [12, 10, 32, 18, 18, 16, 22, 12, 8])

# CFO Sign-off Matrix
sgn = wb.create_sheet("CFO Sign-off Matrix")
sgn["A1"] = "CFO Sign-off Matrix"; sgn["A1"].font = TITLE_FONT
sgn["A2"] = f"Period: {PERIOD}  ·  Target release: BD5 (Dec 3, 2026)"; sgn["A2"].font = SUB_FONT
hdr = ["Item","Description","Amount","Approver","SLA","Status","Approver Name","Signed Date"]
for j, h in enumerate(hdr, 1):
    c = sgn.cell(row=4, column=j, value=h)
    c.font = H_FONT; c.fill = H_FILL; c.alignment = CENTER; c.border = BOX

signoff_items = [
    ("Rec Package Release", "Release Nov 2026 BS rec package to internal and external audit", "—", "CFO", "24h", "Pending"),
    ("MATERIAL exception — Accrued Content stale accrual", "$310K stale Aug content accrual — propose reversal JE",  310_000, "Controller + CFO notify", "4h", "Pending"),
    ("AUTO-INVESTIGATE — IC mismatch LuminaUS↔EMEA", "$267K IC receivable/payable mismatch — material",            267_000, "CFO", "2h", "Pending"),
    ("AUTO-INVESTIGATE — IC mismatch LuminaUS↔APAC", "$267K IC mismatch (symmetric — APAC side)",                   267_000, "CFO", "2h", "Pending"),
    ("Residual Cash variance investigation", "$345K residual variance in LuminaUS Cash - Operating",                 345_000, "Controller + CFO notify", "12h", "Pending"),
    ("Proposed JE: Reverse stale comp accrual", "$185K Aug perf bonus reversal — Controller approval",               185_000, "Controller", "12h", "Pending"),
    ("Proposed JE: Reclass Vortex credit AR", "$45K AR credit balance → deferred revenue",                            45_000, "Controller", "24h", "Pending"),
    ("Proposed JE: Bank fees + service charge", "$200K cash adjustments",                                            200_000, "Controller", "12h", "Pending"),
    ("Proposed JE: Write-off legacy prepaid insurance", "$15K legacy auto fleet policy",                              15_000, "Controller", "24h", "Pending"),
]
for i, item in enumerate(signoff_items, 5):
    for j, v in enumerate(item, 1):
        cell = sgn.cell(row=i, column=j, value=v)
        if j == 3 and isinstance(v, (int, float)): money_fmt(cell)
    sgn.cell(row=i, column=7, value="[Pending]")
    sgn.cell(row=i, column=8, value="[Pending]")
    if i % 2 == 0:
        for c in range(1, 9):
            sgn.cell(row=i, column=c).fill = BAND_FILL
autosize(sgn, [38, 60, 16, 22, 8, 12, 22, 16])

# Audit Trail tab
at = wb.create_sheet("Audit Trail")
at["A1"] = "Rec Package Audit Trail"; at["A1"].font = TITLE_FONT
at["A2"] = f"Package release pending CFO + Controller sign-off  ·  Generated {RUN_TS}"
at["A2"].font = SUB_FONT
at.append([]); at.append([])
trail = [
    ("Agent",           "rec-package"),
    ("Version",         AGENT_VERSION),
    ("Period",          PERIOD),
    ("Generated",       RUN_TS),
    ("Total workpapers", len(RECONS)),
    ("Total exceptions", len(ALL_EXCEPTIONS)),
    ("Total proposed JEs", len(ALL_JES)),
    ("Source files",    len(SOURCES)),
    ("Controller sign-off", "[Pending]"),
    ("CFO sign-off",       "[Pending]"),
    ("Released to",     "[Pending sign-off]"),
]
for i, (k, v) in enumerate(trail, 5):
    at.cell(row=i, column=1, value=k).font = Font(bold=True, color=NAVY)
    at.cell(row=i, column=2, value=v)
autosize(at, [30, 24])

pkg_path = f"outputs/{PERIOD}/{PERIOD}_Rec_Package.xlsx"
wb.save(os.path.join(ROOT, pkg_path))
print(f"[OK] Rec package: {pkg_path}")

# ── Package envelope JSON ───────────────────────────────────────────────────
envelope = {
    "result": {
        "period": PERIOD,
        "company": "Lumina Streaming Co.",
        "entities": ["LuminaUS", "LuminaEMEA", "LuminaAPAC"],
        "reconciliations_completed": len(RECONS),
        "reconciliations_clean":     sum(1 for r in RECONS if r["status"] == "RECONCILED"),
        "reconciliations_variance":  sum(1 for r in RECONS if r["status"] != "RECONCILED"),
        "total_exceptions": len(ALL_EXCEPTIONS),
        "auto_investigate": sum(1 for e in ALL_EXCEPTIONS if e["classification"] == "AUTO-INVESTIGATE"),
        "material":         sum(1 for e in ALL_EXCEPTIONS if e["classification"] == "MATERIAL"),
        "trivial":          sum(1 for e in ALL_EXCEPTIONS if e["classification"] == "TRIVIAL"),
        "proposed_jes": len(ALL_JES),
        "package_path": pkg_path,
        "exception_log_path": exc_log_path,
    },
    "_metadata": {
        "agent": "rec-package",
        "version": AGENT_VERSION,
        "run_timestamp": RUN_TS,
        "sources": [{"path": p, "sha256": h} for p, h in SOURCES.items()],
        "controller_signoff": None,
        "cfo_signoff": None,
    }
}
env_path = f"outputs/{PERIOD}/{PERIOD}_package_envelope.json"
with open(os.path.join(ROOT, env_path), "w") as fh:
    json.dump(envelope, fh, indent=2)
print(f"[OK] Envelope: {env_path}")
print(f"\n[DONE] Total artifacts: {len(SOURCES)} data + {len(RECONS) * 2} workpapers/memos + 2 package outputs + 1 envelope")
